import time
from O365 import Account
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.workbook import Workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

credentials = (
    "5f2bf0de-5d68-4c0a-833f-b32899de4579",
    "Xmv8Q~Kdk4wHu4ARgHOMRx8CVRwPZ.4gqe_4JcRD",
)
account = Account(
    credentials,
    auth_flow_type="credentials",
    tenant_id="d4295056-baca-4708-8c60-0351c0fb01db",
)

if account.authenticate():

    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get(
        "https://security.microsoft.com/securitypoliciesandrules?tid=d4295056-baca-4708-8c60-0351c0fb01db"
    )
    driver.maximize_window()
    email = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.NAME, "loginfmt")))
    email.send_keys("Abhishek@jeeshanahmad2011outlook.onmicrosoft.com")
    email_next_btn = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "idSIButton9")))
    email_next_btn.click()
    pswd = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.NAME, "passwd")))
    pswd.send_keys("12345@Fdsa")
    time.sleep(2)
    signIn = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "idSIButton9")))
    signIn.click()
    loginIn = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "idSIButton9")))
    loginIn.click()
    time.sleep(2)
    back_to_all_policy = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
        (By.XPATH, "//a[normalize-space()='Threat policies']")))
    back_to_all_policy.click()

    ############ firstPolicy#############

    policy_number_first = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
        (By.XPATH, "//button[normalize-space()='Anti-phishing']")))
    policy_number_first.click()
    policy_number_first_data = WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "ms-DetailsRow-fields")))
    wb = Workbook()
    FIRST_FILE = "./all_policy/Anti-phishing.xlsx"
    for data in policy_number_first_data:
        data.click()
        policy_first_data = WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "ms-MetaDataItem")))
        ws1 = wb.create_sheet(data.text.splitlines()[0])
        for data in policy_first_data:
            sav_data = [data.text.splitlines()]
            for row in sav_data:
                ws1.append(row)
        close_btn = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Close')]")))
        close_btn.click()
    wb.save(filename=FIRST_FILE)
    driver.back()

    ############ secondPolicy#############

    inside_second_Policy_Folder = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//button[normalize-space()='Anti-spam']")))
    inside_second_Policy_Folder.click()
    second_Policy_Item = WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "ms-DetailsRow-fields")))
    wb = Workbook()
    SECOND_FILE = "./all_policy/Anti-spam.xlsx"
    for data in second_Policy_Item:
        ws1 = wb.create_sheet(data.text.splitlines()[0])
        data.click()
        all_items = WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "ms-MetaDataItem")))
        semi_final = []
        for res in all_items:
            second_policy_result = res.text.splitlines()
            semi_final.append(second_policy_result)
            sav_data = [res.text.splitlines()]
            for row in sav_data:
                ws1.append(row)
        wb.save(filename=SECOND_FILE)
        try:
            button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                (By.XPATH, "//button[@aria-label='Edit allowed and blocked senders and domains']")))
            if button.is_displayed():
                driver.execute_script("arguments[0].click();", button)
                get_data = WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located(
                    (By.XPATH, "//div[@id='fluent-default-layer-host']/div/div/div/div/div[2]/div[2]/div/div[4]/div/button")))
                time.sleep(2)
                for i in get_data:
                    i.click()
                    items_ = WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located(
                        (By.CLASS_NAME, "ms-DetailsList-contentWrapper")))
                    for new_res in items_:
                        second_policy_ = new_res.text.splitlines()
                    semi_final.append(second_policy_)
                    save_data = [i.text.splitlines(), second_policy_]
                    A = "\uea3a"
                    try:
                        for A in second_policy_:
                            second_policy_.remove("\uea3a")
                    except:
                        print()
                    flat_list = [
                        item for sublist in save_data for item in sublist]
                    print(flat_list)
                    ws1.append(flat_list)
                    wb.save(filename=SECOND_FILE)
                    semi_final.append(second_policy_)
                    done_btn = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Done']")))
                    done_btn.click()
                cancel_btn = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//i[@data-icon-name='Cancel']")))
                cancel_btn.click()
            close_btn = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                (By.XPATH, "//span[contains(text(),'Close')]]")))
            close_btn.click()
        except:
            close_btn = driver.find_element(
                By.XPATH, "//span[contains(text(),'Close')]"
            ).click()
    driver.back()

    # ########### thirdPolicy################

    THIRD_FILE = "./all_policy/Anti-malware.xlsx"
    policy_third = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
        (By.XPATH, "//button[normalize-space()='Anti-malware']")))
    policy_third.click()
    time.sleep(4)
    data_policy_third_number_policy = WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "ms-DetailsRow-fields")))
    wb = Workbook()
    for third_data in data_policy_third_number_policy:
        ws1 = wb.create_sheet(third_data.text.splitlines()[0])
        third_data.click()
        data_policy_third_number = WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "ms-MetaDataItem")))
        for data in data_policy_third_number:
            sav_data = [data.text.splitlines()]
            for row in sav_data:
                ws1.append(row)
        close_btn = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Close')]")))
        close_btn.click()
    wb.save(filename=THIRD_FILE)
    driver.back()
    secure_Score = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//a[@name='Secure score']")))
    secure_Score.click()
    recomm_action = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
        (By.XPATH, "//button[@aria-label='Recommended actions']")))
    recomm_action.click()
    secure_Score = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Export')]")))
    secure_Score.click()
    driver.quit()